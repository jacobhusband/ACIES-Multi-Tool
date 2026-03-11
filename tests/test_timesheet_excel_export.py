import base64
import math
import posixpath
import sys
import tempfile
import types
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from unittest.mock import patch


def _ensure_google_genai_stub():
    try:
        from google import genai as _genai  # noqa: F401
        from google.genai import types as _types  # noqa: F401
        return
    except Exception:
        google_module = sys.modules.get("google")
        if google_module is None:
            google_module = types.ModuleType("google")
            google_module.__path__ = []
            sys.modules["google"] = google_module

        genai_module = types.ModuleType("google.genai")
        genai_types_module = types.ModuleType("google.genai.types")
        genai_module.types = genai_types_module
        google_module.genai = genai_module

        sys.modules["google.genai"] = genai_module
        sys.modules["google.genai.types"] = genai_types_module


def _ensure_webview_stub():
    try:
        import webview  # noqa: F401
        return
    except Exception:
        webview_module = types.ModuleType("webview")
        webview_module.windows = []
        webview_module.create_window = lambda *args, **kwargs: None
        webview_module.start = lambda *args, **kwargs: None
        sys.modules["webview"] = webview_module


def _ensure_dotenv_stub():
    try:
        from dotenv import load_dotenv as _load_dotenv  # noqa: F401
        return
    except Exception:
        dotenv_module = types.ModuleType("dotenv")
        dotenv_module.load_dotenv = lambda *args, **kwargs: False
        sys.modules["dotenv"] = dotenv_module


def _ensure_requests_stub():
    try:
        import requests  # noqa: F401
        return
    except Exception:
        requests_module = types.ModuleType("requests")
        requests_module.get = lambda *args, **kwargs: None
        sys.modules["requests"] = requests_module


def _ensure_pydantic_stub():
    try:
        from pydantic import BaseModel as _BaseModel, Field as _Field  # noqa: F401
        return
    except Exception:
        pydantic_module = types.ModuleType("pydantic")

        class BaseModel:
            pass

        def Field(*args, **kwargs):
            if args:
                return args[0]
            return kwargs.get("default")

        pydantic_module.BaseModel = BaseModel
        pydantic_module.Field = Field
        sys.modules["pydantic"] = pydantic_module


_ensure_google_genai_stub()
_ensure_webview_stub()
_ensure_dotenv_stub()
_ensure_requests_stub()
_ensure_pydantic_stub()

import main as main_module
from main import Api
from openpyxl import load_workbook
from openpyxl.utils.units import EMU_to_pixels, points_to_pixels
from PIL import Image as PILImage


REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = REPO_ROOT / "templates" / "Template_Timesheet.xlsx"
EXPENSE_SHEET_NAME = "Project Expense Sheet"
DAY_TOTALS = {day: 0 for day in ("mon", "tue", "wed", "thu", "fri", "sat", "sun")}
PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR4nGNgYAAAAAMAASsJTYQAAAAASUVORK5CYII="
)
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "docrel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
}


def _write_image(path, size, image_format="PNG", orientation=None, color=(220, 64, 64)):
    image = PILImage.new("RGB", size, color)
    save_kwargs = {"format": image_format}
    if orientation is not None:
        exif = image.getexif()
        exif[274] = orientation
        save_kwargs["exif"] = exif.tobytes()
    image.save(path, **save_kwargs)


def _make_expense_entry(index, mileage=None, expense=None):
    value = index + 1
    return {
        "date": f"2026-03-{10 + index:02d}",
        "description": f"Expense {value}",
        "mileage": value if mileage is None else mileage,
        "expense": (value * 10) if expense is None else expense,
    }


def _make_project(name, project_id, entry_count, image_paths=None):
    return {
        "projectName": name,
        "projectId": project_id,
        "entries": [_make_expense_entry(index) for index in range(entry_count)],
        "images": [{"path": path} for path in (image_paths or [])],
    }


def _make_timesheet_entry(project_id, project_name, task_number, hours=None, mileage=0):
    return {
        "projectId": project_id,
        "taskNumber": task_number,
        "projectName": project_name,
        "function": "E",
        "pmInitials": "JH",
        "serviceDescription": f"Work for {project_name}",
        "hours": hours or {},
        "mileage": mileage,
    }


class TimesheetExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.api = Api.__new__(Api)

    def _export_timesheet_workbook(self, projects, output_path, entries=None):
        payload = {
            "weekKey": "2026-03-09",
            "weekDisplay": "Week of 03/09/26",
            "userName": "Tester",
            "filePath": str(output_path),
            "entries": entries or [],
            "totals": dict(DAY_TOTALS, mileage=0),
            "expenses": {
                "projects": projects,
                "mileageRate": 0.70,
            },
        }
        with patch.object(main_module.os, "startfile", create=True):
            result = self.api.export_timesheet_excel(payload)
        self.assertEqual("success", result["status"])
        self.assertTrue(output_path.exists())
        return output_path

    def _export_expense_workbook(self, projects, week_key):
        with tempfile.TemporaryDirectory(prefix="expense-home-") as temp_dir:
            home_dir = Path(temp_dir)
            documents_dir = home_dir / "Documents"
            documents_dir.mkdir(parents=True, exist_ok=True)
            with patch.object(main_module.os.path, "expanduser", return_value=str(home_dir)):
                with patch.object(main_module.os, "startfile", create=True):
                    result = self.api.export_expense_sheet_excel({
                        "weekKey": week_key,
                        "projects": projects,
                        "mileageRate": 0.70,
                    })
            self.assertEqual("success", result["status"])
            output_path = documents_dir / f"Expense_Sheet_{week_key}.xlsx"
            self.assertTrue(output_path.exists())
            final_path = Path(tempfile.gettempdir()) / f"{week_key}.xlsx"
            final_path.write_bytes(output_path.read_bytes())
        return final_path

    def _load_expense_sheet(self, workbook_path):
        workbook = load_workbook(workbook_path)
        template = load_workbook(TEMPLATE_PATH)
        self.addCleanup(workbook.close)
        self.addCleanup(template.close)
        return workbook[EXPENSE_SHEET_NAME], template[EXPENSE_SHEET_NAME]

    def _assert_style_matches(self, worksheet, actual_coord, template_sheet, template_coord):
        actual = worksheet[actual_coord]
        expected = template_sheet[template_coord]
        self.assertEqual(expected.style_id, actual.style_id, actual_coord)
        self.assertEqual(expected.number_format, actual.number_format, actual_coord)
        self._assert_border_matches(worksheet, actual_coord, template_sheet, template_coord)

    def _assert_border_matches(self, worksheet, actual_coord, template_sheet, template_coord):
        actual = worksheet[actual_coord]
        expected = template_sheet[template_coord]
        self.assertEqual(expected.border.left.style, actual.border.left.style, actual_coord)
        self.assertEqual(expected.border.right.style, actual.border.right.style, actual_coord)
        self.assertEqual(expected.border.top.style, actual.border.top.style, actual_coord)
        self.assertEqual(expected.border.bottom.style, actual.border.bottom.style, actual_coord)

    def _assert_merged(self, worksheet, range_ref):
        merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
        self.assertIn(range_ref, merged_ranges)

    def _get_first_image_anchor_row(self, workbook_path, sheet_name):
        anchors = self._get_image_anchors(workbook_path, sheet_name)
        return anchors[0]["row"]

    def _get_image_anchors(self, workbook_path, sheet_name):
        with zipfile.ZipFile(workbook_path) as archive:
            workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
            workbook_rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            workbook_rels = {
                rel.attrib["Id"]: rel.attrib["Target"]
                for rel in workbook_rels_xml.findall("pkgrel:Relationship", XML_NS)
            }

            sheet_path = None
            for sheet in workbook_xml.findall("main:sheets/main:sheet", XML_NS):
                if sheet.attrib["name"] != sheet_name:
                    continue
                rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                sheet_target = workbook_rels[rel_id]
                normalized_target = sheet_target.lstrip("/")
                if normalized_target.startswith("xl/"):
                    sheet_path = posixpath.normpath(normalized_target)
                else:
                    sheet_path = posixpath.normpath(posixpath.join("xl", normalized_target))
                break

            self.assertIsNotNone(sheet_path)

            sheet_rel_path = posixpath.join(
                posixpath.dirname(sheet_path),
                "_rels",
                f"{posixpath.basename(sheet_path)}.rels",
            )
            sheet_rels_xml = ET.fromstring(archive.read(sheet_rel_path))
            drawing_target = None
            for rel in sheet_rels_xml.findall("pkgrel:Relationship", XML_NS):
                if rel.attrib["Type"].endswith("/drawing"):
                    drawing_target = rel.attrib["Target"]
                    break

            self.assertIsNotNone(drawing_target)
            normalized_drawing_target = drawing_target.lstrip("/")
            if normalized_drawing_target.startswith("xl/"):
                drawing_path = posixpath.normpath(normalized_drawing_target)
            else:
                drawing_path = posixpath.normpath(
                    posixpath.join(posixpath.dirname(sheet_path), normalized_drawing_target)
                )
            drawing_xml = ET.fromstring(archive.read(drawing_path))

            anchors = []
            for anchor_node in drawing_xml.findall("xdr:oneCellAnchor", XML_NS):
                row_node = anchor_node.find("xdr:from/xdr:row", XML_NS)
                col_node = anchor_node.find("xdr:from/xdr:col", XML_NS)
                ext_node = anchor_node.find("xdr:ext", XML_NS)
                self.assertIsNotNone(row_node)
                self.assertIsNotNone(ext_node)
                anchors.append({
                    "row": int(row_node.text) + 1,
                    "col": int(col_node.text) + 1 if col_node is not None else None,
                    "width_pixels": EMU_to_pixels(int(ext_node.attrib["cx"])),
                    "height_pixels": EMU_to_pixels(int(ext_node.attrib["cy"])),
                })

            if not anchors:
                for anchor_node in drawing_xml.findall("xdr:twoCellAnchor", XML_NS):
                    row_node = anchor_node.find("xdr:from/xdr:row", XML_NS)
                    col_node = anchor_node.find("xdr:from/xdr:col", XML_NS)
                    self.assertIsNotNone(row_node)
                    anchors.append({
                        "row": int(row_node.text) + 1,
                        "col": int(col_node.text) + 1 if col_node is not None else None,
                        "width_pixels": None,
                        "height_pixels": None,
                    })

            self.assertTrue(anchors)
            return anchors

    def _excel_column_width_to_pixels(self, width):
        return int(math.floor(((256 * float(width) + math.floor(128 / 7)) / 256) * 7))

    def _get_total_column_width_pixels(self, worksheet, start_column=1, end_column=5):
        total_width = 0
        default_width = worksheet.sheet_format.defaultColWidth or 8.43
        for column_index in range(start_column, end_column + 1):
            column_letter = chr(ord("A") + column_index - 1)
            column_width = worksheet.column_dimensions[column_letter].width
            total_width += self._excel_column_width_to_pixels(
                default_width if column_width is None else column_width
            )
        return total_width

    def _get_row_span_for_height(self, worksheet, start_row, height_pixels):
        covered_height = 0
        row_span = 0
        next_row = start_row
        while covered_height < height_pixels:
            row_height = worksheet.row_dimensions[next_row].height
            if row_height is None:
                row_height = worksheet.sheet_format.defaultRowHeight or 15
            covered_height += max(points_to_pixels(row_height), 1)
            row_span += 1
            next_row += 1
        return row_span

    def test_combined_export_preserves_second_project_section_formatting(self):
        projects = [
            _make_project("Alpha", "1001", 1),
            _make_project("Beta", "1002", 1),
        ]

        with tempfile.TemporaryDirectory(prefix="timesheet-export-") as temp_dir:
            output_path = Path(temp_dir) / "combined.xlsx"
            workbook_path = self._export_timesheet_workbook(projects, output_path)
            worksheet, template_sheet = self._load_expense_sheet(workbook_path)

        self.assertEqual("PROJECT: Beta", worksheet["A13"].value)
        self.assertEqual("JOB #: 1002", worksheet["A14"].value)
        self._assert_style_matches(worksheet, "A13", template_sheet, "A1")
        self._assert_style_matches(worksheet, "B15", template_sheet, "B3")
        self._assert_style_matches(worksheet, "C15", template_sheet, "C3")
        self._assert_style_matches(worksheet, "A16", template_sheet, "A4")
        self._assert_style_matches(worksheet, "B16", template_sheet, "B4")
        self._assert_style_matches(worksheet, "D21", template_sheet, "D9")
        self._assert_style_matches(worksheet, "C23", template_sheet, "C11")
        self.assertEqual(template_sheet.row_dimensions[3].height, worksheet.row_dimensions[15].height)
        self.assertEqual(template_sheet.row_dimensions[4].height, worksheet.row_dimensions[16].height)
        for range_ref in ("B15:C15", "B16:C16", "B17:C17", "B18:C18", "B19:C19", "B20:C20", "D23:E23"):
            self._assert_merged(worksheet, range_ref)

    def test_overflow_rows_keep_formatting_and_shift_later_sections(self):
        projects = [
            _make_project("Alpha", "2001", 7),
            _make_project("Beta", "2002", 1),
        ]

        with tempfile.TemporaryDirectory(prefix="timesheet-overflow-") as temp_dir:
            output_path = Path(temp_dir) / "overflow.xlsx"
            workbook_path = self._export_timesheet_workbook(projects, output_path)
            worksheet, template_sheet = self._load_expense_sheet(workbook_path)

        self.assertEqual("PROJECT: Beta", worksheet["A15"].value)
        self.assertEqual("JOB #: 2002", worksheet["A16"].value)
        self._assert_style_matches(worksheet, "A9", template_sheet, "A5")
        self._assert_style_matches(worksheet, "B9", template_sheet, "B5")
        self._assert_style_matches(worksheet, "C9", template_sheet, "C5")
        self._assert_style_matches(worksheet, "D9", template_sheet, "D5")
        self._assert_border_matches(worksheet, "E9", template_sheet, "E5")
        self._assert_style_matches(worksheet, "A10", template_sheet, "A5")
        self.assertEqual(template_sheet.row_dimensions[5].height, worksheet.row_dimensions[9].height)
        self.assertEqual(template_sheet.row_dimensions[5].height, worksheet.row_dimensions[10].height)
        self._assert_merged(worksheet, "B9:C9")
        self._assert_merged(worksheet, "B10:C10")
        self._assert_merged(worksheet, "D13:E13")
        self.assertEqual("Expense 7", worksheet["B10"].value)
        self.assertEqual("=SUM(D4:D10)", worksheet["D11"].value)
        self.assertEqual("=SUM(E4:E10)", worksheet["E11"].value)
        self._assert_style_matches(worksheet, "A15", template_sheet, "A1")
        self._assert_merged(worksheet, "B17:C17")
        self._assert_merged(worksheet, "D25:E25")
        self.assertEqual("Expense 1", worksheet["B18"].value)

    def test_expense_only_export_uses_same_section_renderer(self):
        projects = [
            _make_project("Gamma", "3001", 1),
            _make_project("Delta", "3002", 1),
        ]

        workbook_path = self._export_expense_workbook(projects, "expense-shared-renderer")
        worksheet, template_sheet = self._load_expense_sheet(workbook_path)

        self.assertEqual("PROJECT: Delta", worksheet["A13"].value)
        self._assert_style_matches(worksheet, "A13", template_sheet, "A1")
        self._assert_style_matches(worksheet, "B15", template_sheet, "B3")
        self._assert_style_matches(worksheet, "A16", template_sheet, "A4")
        self._assert_merged(worksheet, "B15:C15")
        self._assert_merged(worksheet, "D23:E23")

    def test_expense_cells_keep_currency_formatting(self):
        projects = [
            _make_project("Money Job 1", "3500", 2),
            _make_project("Money Job 2", "3501", 1),
            _make_project("Money Job 3", "3502", 1),
        ]

        with tempfile.TemporaryDirectory(prefix="timesheet-currency-") as temp_dir:
            output_path = Path(temp_dir) / "currency.xlsx"
            workbook_path = self._export_timesheet_workbook(projects, output_path)
            worksheet, template_sheet = self._load_expense_sheet(workbook_path)

        currency_format = template_sheet["E4"].number_format
        subtotal_currency_format = template_sheet["E9"].number_format
        rate_currency_format = template_sheet["D10"].number_format
        total_currency_format = template_sheet["D11"].number_format

        self.assertEqual(10, worksheet["E4"].value)
        self.assertEqual(20, worksheet["E5"].value)
        self.assertEqual(10, worksheet["E16"].value)
        self.assertEqual(10, worksheet["E28"].value)
        self.assertEqual('"$"#,##0.00', worksheet["E4"].number_format)
        self.assertEqual(currency_format, worksheet["E4"].number_format)
        self.assertEqual(currency_format, worksheet["E5"].number_format)
        self.assertEqual(currency_format, worksheet["E16"].number_format)
        self.assertEqual(currency_format, worksheet["E17"].number_format)
        self.assertEqual(currency_format, worksheet["E28"].number_format)
        self.assertEqual(currency_format, worksheet["E29"].number_format)
        self.assertEqual(subtotal_currency_format, worksheet["E9"].number_format)
        self.assertEqual(subtotal_currency_format, worksheet["E21"].number_format)
        self.assertEqual(subtotal_currency_format, worksheet["E33"].number_format)
        self.assertEqual(rate_currency_format, worksheet["D10"].number_format)
        self.assertEqual(rate_currency_format, worksheet["D22"].number_format)
        self.assertEqual(rate_currency_format, worksheet["D34"].number_format)
        self.assertEqual(total_currency_format, worksheet["D11"].number_format)
        self.assertEqual(total_currency_format, worksheet["D23"].number_format)
        self.assertEqual(total_currency_format, worksheet["D35"].number_format)
        self.assertEqual("=SUM(E4:E8)", worksheet["E9"].value)
        self.assertEqual("=SUM(E16:E20)", worksheet["E21"].value)
        self.assertEqual("=SUM(E28:E32)", worksheet["E33"].value)
        self.assertEqual("=D10+E9", worksheet["D11"].value)
        self.assertEqual("=D22+E21", worksheet["D23"].value)
        self.assertEqual("=D34+E33", worksheet["D35"].value)

    def test_images_anchor_below_shifted_signature_block(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-images-") as temp_dir:
            image_path = Path(temp_dir) / "receipt.png"
            image_path.write_bytes(PNG_BYTES)
            projects = [
                _make_project("One", "4001", 1, [str(image_path)]),
                _make_project("Two", "4002", 1),
                _make_project("Three", "4003", 1),
                _make_project("Four", "4004", 1),
            ]

            output_path = Path(temp_dir) / "images.xlsx"
            workbook_path = self._export_timesheet_workbook(projects, output_path)
            worksheet, template_sheet = self._load_expense_sheet(workbook_path)
            anchor_row = self._get_first_image_anchor_row(workbook_path, EXPENSE_SHEET_NAME)

        self.assertEqual("PROJECT: Four", worksheet["A39"].value)
        self._assert_style_matches(worksheet, "A39", template_sheet, "A1")
        self.assertEqual(57, anchor_row)

    def test_images_export_from_every_project_in_order(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-order-") as temp_dir:
            temp_dir = Path(temp_dir)
            image_sizes = [(120, 60), (100, 100), (60, 120), (200, 100)]
            image_paths = []
            for index, size in enumerate(image_sizes, start=1):
                image_path = temp_dir / f"receipt-{index}.png"
                _write_image(image_path, size)
                image_paths.append(str(image_path))

            projects = [
                _make_project("One", "4101", 1, [image_paths[0]]),
                _make_project("Two", "4102", 1, [image_paths[1], image_paths[2]]),
                _make_project("Three", "4103", 1, [image_paths[3]]),
            ]
            output_path = temp_dir / "ordered-images.xlsx"
            workbook_path = self._export_timesheet_workbook(projects, output_path)
            worksheet, _template_sheet = self._load_expense_sheet(workbook_path)
            anchors = self._get_image_anchors(workbook_path, EXPENSE_SHEET_NAME)
            target_width = self._get_total_column_width_pixels(worksheet)

        self.assertEqual(4, len(anchors))
        self.assertEqual(sorted(anchor["row"] for anchor in anchors), [anchor["row"] for anchor in anchors])
        expected_heights = [
            round(target_width * (60 / 120)),
            round(target_width * (100 / 100)),
            round(target_width * (120 / 60)),
            round(target_width * (100 / 200)),
        ]
        for anchor in anchors:
            self.assertAlmostEqual(target_width, anchor["width_pixels"], delta=1)
        for anchor, expected_height in zip(anchors, expected_heights):
            self.assertAlmostEqual(expected_height, anchor["height_pixels"], delta=1)

    def test_exported_images_match_columns_a_through_e_width(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-width-") as temp_dir:
            temp_dir = Path(temp_dir)
            image_path = temp_dir / "wide.png"
            _write_image(image_path, (240, 120))
            output_path = temp_dir / "image-width.xlsx"
            workbook_path = self._export_timesheet_workbook(
                [_make_project("Width", "4201", 1, [str(image_path)])],
                output_path,
            )
            worksheet, _template_sheet = self._load_expense_sheet(workbook_path)
            anchor = self._get_image_anchors(workbook_path, EXPENSE_SHEET_NAME)[0]
            target_width = self._get_total_column_width_pixels(worksheet)

        self.assertAlmostEqual(target_width, anchor["width_pixels"], delta=1)
        self.assertAlmostEqual(round(target_width * 0.5), anchor["height_pixels"], delta=1)

    def test_exported_images_stack_without_overlap(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-stack-") as temp_dir:
            temp_dir = Path(temp_dir)
            first_path = temp_dir / "first.png"
            second_path = temp_dir / "second.png"
            _write_image(first_path, (150, 150))
            _write_image(second_path, (150, 150), color=(64, 64, 220))
            output_path = temp_dir / "stacked-images.xlsx"
            workbook_path = self._export_timesheet_workbook(
                [_make_project("Stack", "4301", 1, [str(first_path), str(second_path)])],
                output_path,
            )
            worksheet, _template_sheet = self._load_expense_sheet(workbook_path)
            anchors = self._get_image_anchors(workbook_path, EXPENSE_SHEET_NAME)

        self.assertEqual(2, len(anchors))
        expected_second_row = (
            anchors[0]["row"]
            + self._get_row_span_for_height(worksheet, anchors[0]["row"], anchors[0]["height_pixels"])
            + 1
        )
        self.assertEqual(expected_second_row, anchors[1]["row"])

    def test_exported_images_respect_exif_orientation(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-orientation-") as temp_dir:
            temp_dir = Path(temp_dir)
            image_path = temp_dir / "oriented.jpg"
            _write_image(image_path, (120, 60), image_format="JPEG", orientation=6)
            output_path = temp_dir / "oriented-image.xlsx"
            workbook_path = self._export_timesheet_workbook(
                [_make_project("Orientation", "4401", 1, [str(image_path)])],
                output_path,
            )
            worksheet, _template_sheet = self._load_expense_sheet(workbook_path)
            anchor = self._get_image_anchors(workbook_path, EXPENSE_SHEET_NAME)[0]
            target_width = self._get_total_column_width_pixels(worksheet)

        self.assertAlmostEqual(target_width, anchor["width_pixels"], delta=1)
        self.assertGreater(anchor["height_pixels"], anchor["width_pixels"])
        self.assertAlmostEqual(round(target_width * 2), anchor["height_pixels"], delta=2)

    def test_expense_image_preview_returns_data_url(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-preview-") as temp_dir:
            temp_dir = Path(temp_dir)
            image_path = temp_dir / "preview.png"
            _write_image(image_path, (180, 120))
            result = self.api.get_expense_image_preview(str(image_path), max_size=90)

        self.assertEqual("success", result["status"])
        self.assertTrue(result["dataUrl"].startswith("data:image/png;base64,"))
        self.assertLessEqual(max(result["width"], result["height"]), 90)
        self.assertEqual(str(image_path), result["path"])

    def test_expense_image_preview_resolves_moved_attachment_path(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-preview-moved-") as temp_dir:
            temp_dir = Path(temp_dir)
            desktop_root = temp_dir / "Desktop"
            moved_dir = desktop_root / "Expenses" / "BAC Kent, WA Expenses"
            moved_dir.mkdir(parents=True, exist_ok=True)
            moved_image_path = moved_dir / "IMG_0466.JPG"
            _write_image(moved_image_path, (160, 120), image_format="JPEG")

            stale_path = desktop_root / "BAC Kent, WA Expenses" / "IMG_0466.JPG"
            result = self.api.get_expense_image_preview(str(stale_path), max_size=120)

        self.assertEqual("success", result["status"])
        self.assertEqual(str(moved_image_path), result["path"])
        self.assertTrue(result["dataUrl"].startswith("data:image/png;base64,"))

    def test_export_uses_moved_attachment_when_stored_path_is_stale(self):
        with tempfile.TemporaryDirectory(prefix="timesheet-image-relocate-") as temp_dir:
            temp_dir = Path(temp_dir)
            desktop_root = temp_dir / "Desktop"
            moved_dir = desktop_root / "Expenses" / "BAC Kent, WA Expenses"
            moved_dir.mkdir(parents=True, exist_ok=True)
            moved_image_path = moved_dir / "IMG_0466.JPG"
            _write_image(moved_image_path, (160, 120), image_format="JPEG")

            stale_path = desktop_root / "BAC Kent, WA Expenses" / "IMG_0466.JPG"
            output_path = temp_dir / "moved-image.xlsx"
            workbook_path = self._export_timesheet_workbook(
                [_make_project("Moved", "4501", 1, [str(stale_path)])],
                output_path,
            )
            worksheet, _template_sheet = self._load_expense_sheet(workbook_path)
            anchors = self._get_image_anchors(workbook_path, EXPENSE_SHEET_NAME)
            target_width = self._get_total_column_width_pixels(worksheet)

        self.assertEqual(1, len(anchors))
        self.assertAlmostEqual(target_width, anchors[0]["width_pixels"], delta=1)

    def test_time_log_totals_use_sum_formulas(self):
        entries = [
            _make_timesheet_entry(
                "5001",
                "Alpha",
                "10",
                hours={"mon": 1.5, "tue": 2, "wed": 3, "thu": 4, "fri": 5, "sat": 6, "sun": 7},
                mileage=8,
            ),
            _make_timesheet_entry(
                "5002",
                "Beta",
                "20",
                hours={"mon": 0.5, "wed": 1.25, "fri": 2.75},
                mileage=4,
            ),
        ]

        with tempfile.TemporaryDirectory(prefix="timesheet-formulas-") as temp_dir:
            output_path = Path(temp_dir) / "timelog.xlsx"
            workbook_path = self._export_timesheet_workbook([], output_path, entries=entries)
            workbook = load_workbook(workbook_path)
            self.addCleanup(workbook.close)
            worksheet = workbook["time log"]

        self.assertEqual("=SUM(K5:K25)", worksheet["K26"].value)
        self.assertEqual("=SUM(L5:L25)", worksheet["L26"].value)
        self.assertEqual("=SUM(M5:M25)", worksheet["M26"].value)
        self.assertEqual("=SUM(N5:N25)", worksheet["N26"].value)
        self.assertEqual("=SUM(O5:O25)", worksheet["O26"].value)
        self.assertEqual("=SUM(P5:P25)", worksheet["P26"].value)
        self.assertEqual("=SUM(Q5:Q25)", worksheet["Q26"].value)
        self.assertEqual("=SUM(R5:R25)", worksheet["R26"].value)


if __name__ == "__main__":
    unittest.main()
